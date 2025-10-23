import random
import heapq
import statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

print("=" * 70)
print("OFFICE PRINT QUEUE SIMULATION")
print("=" * 70)

class PrintJob:
    def __init__(self, arrival_time):
        self.arrival_time = arrival_time
        self.start_time = None
        self.departure_time = None

def generate_arrivals(arrival_rate, sim_time, seed=42):
    rng = random.Random(seed)
    arrivals = []
    time = 0.0
    while True:
        inter = rng.expovariate(arrival_rate)
        time += inter
        if time >= sim_time:
            break
        arrivals.append(time)
    return arrivals

def generate_services(num_jobs, service_rate, seed=43):
    rng = random.Random(seed)
    return [rng.expovariate(service_rate) for _ in range(num_jobs)]

def simulate_print_queue(num_printers, arrival_rate, service_rate, sim_time, arr_seed=42, serv_seed=43):
    arrivals = generate_arrivals(arrival_rate, sim_time, arr_seed)
    services = generate_services(len(arrivals), service_rate, serv_seed)
    service_iter = iter(services)
    
    event_queue = []
    jobs = []
    printers = [False] * num_printers
    queue = []
    
    for arr_time in arrivals:
        heapq.heappush(event_queue, (arr_time, 'arrival', None))
    
    env_time = 0.0
    while event_queue:
        time, event_type, data = heapq.heappop(event_queue)
        if time > sim_time:
            break
        env_time = time
        
        if event_type == 'arrival':
            job = PrintJob(env_time)
            jobs.append(job)
            idle_printer = next((i for i, busy in enumerate(printers) if not busy), None)
            if idle_printer is not None:
                printers[idle_printer] = True
                job.start_time = env_time
                service_time = next(service_iter)
                departure_time = env_time + service_time
                job.departure_time = departure_time
                heapq.heappush(event_queue, (departure_time, 'departure', idle_printer))
            else:
                queue.append(job)
        
        elif event_type == 'departure':
            printer_id = data
            printers[printer_id] = False
            if queue:
                job = queue.pop(0)
                job.start_time = env_time
                service_time = next(service_iter)
                departure_time = env_time + service_time
                job.departure_time = departure_time
                printers[printer_id] = True
                heapq.heappush(event_queue, (departure_time, 'departure', printer_id))
    
    completed_jobs = [j for j in jobs if j.departure_time is not None and j.departure_time <= sim_time]
    if not completed_jobs:
        return {
            'avg_wait_time': 0, 'max_wait_time': 0, 'avg_queue_length': 0,
            'throughput': 0, 'utilization': 0, 'num_jobs': len(jobs), 'num_completed': 0
        }
    
    wait_times = [j.start_time - j.arrival_time for j in completed_jobs]
    avg_wait_time = statistics.mean(wait_times)
    
    total_wait_time = sum(wait_times)
    avg_queue_length = total_wait_time / sim_time
    
    throughput = len(completed_jobs) / sim_time
    
    service_times = [j.departure_time - j.start_time for j in completed_jobs]
    total_service_time = sum(service_times)
    utilization = total_service_time / (sim_time * num_printers)
    
    return {
        'avg_wait_time': round(avg_wait_time, 3),
        'max_wait_time': round(max(wait_times), 3),
        'avg_queue_length': round(avg_queue_length, 3),
        'throughput': round(throughput, 3),
        'utilization': round(utilization * 100, 1),
        'num_jobs': len(jobs),
        'num_completed': len(completed_jobs)
    }

def print_results(scenario_name, results):
    print(f"\n{scenario_name}")
    print("-" * 40)
    print(f"Avg Wait Time:        {results['avg_wait_time']} minutes")
    print(f"Max Wait Time:        {results['max_wait_time']} minutes")
    print(f"Avg Queue Length:     {results['avg_queue_length']} jobs")
    print(f"Throughput:           {results['throughput']} jobs/min")
    print(f"Utilization:          {results['utilization']}%")
    print(f"Total Jobs:           {results['num_jobs']}")
    print(f"Completed Jobs:       {results['num_completed']}")

def save_to_txt(scenario_name, results):
    """Save results to text file"""
    with open("results.txt", "a") as f:
        f.write(f"\n{scenario_name}\n")
        f.write("-" * 40 + "\n")
        for key, value in results.items():
            f.write(f"{key.replace('_', ' ').title()}: {value}\n")
        f.write("\n")

def create_excel_data_only(results1, results2, results3):
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Print Queue Results"
    
    # Data for table
    data = [
        ['Metric', 'Base Case (1 Printer)', '2 Printers', 'High Traffic'],
        ['Avg Wait Time (min)', results1['avg_wait_time'], results2['avg_wait_time'], results3['avg_wait_time']],
        ['Max Wait Time (min)', results1['max_wait_time'], results2['max_wait_time'], results3['max_wait_time']],
        ['Avg Queue Length', results1['avg_queue_length'], results2['avg_queue_length'], results3['avg_queue_length']],
        ['Throughput (jobs/min)', results1['throughput'], results2['throughput'], results3['throughput']],
        ['Utilization (%)', results1['utilization'], results2['utilization'], results3['utilization']],
        ['Total Jobs', results1['num_jobs'], results2['num_jobs'], results3['num_jobs']],
        ['Completed Jobs', results1['num_completed'], results2['num_completed'], results3['num_completed']]
    ]
    
    # Write data to Excel
    for row in data:
        ws.append(row)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save Excel file
    filename = "results.xlsx"
    wb.save(filename)
    print(f"Excel file saved: {filename}")
    return filename

#RUN ALL 3 SCENARIOS
SIM_TIME = 480

with open("results.txt", "w") as f:
    f.write("OFFICE PRINT QUEUE SIMULATION RESULTS\n")
    f.write("=" * 50 + "\n\n")

print("\nSCENARIO 1: BASE CASE (1 Printer, Moderate Traffic)")
results1 = simulate_print_queue(1, 0.2, 0.5, SIM_TIME)
print_results("SCENARIO 1: BASE CASE", results1)
save_to_txt("SCENARIO 1: BASE CASE", results1)

print("\nSCENARIO 2: ADD SECOND PRINTER") 
results2 = simulate_print_queue(2, 0.2, 0.5, SIM_TIME)
print_results("SCENARIO 2: 2 PRINTERS", results2)
save_to_txt("SCENARIO 2: 2 PRINTERS", results2)

print("\nSCENARIO 3: HIGH TRAFFIC (More Jobs)")
results3 = simulate_print_queue(1, 0.3, 0.5, SIM_TIME)
print_results("SCENARIO 3: HIGH TRAFFIC", results3)
save_to_txt("SCENARIO 3: HIGH TRAFFIC", results3)

#CREATE EXCEL SHEET
excel_file = create_excel_data_only(results1, results2, results3)

#FINAL SUMMARY
print("\n" + "=" * 70)
print("FINAL COMPARISON SUMMARY")
print("=" * 70)
print(f"{'Metric':<20} {'Base':<10} {'2 Printers':<12} {'High Traffic':<12}")
print("-" * 70)
print(f"Avg Wait Time    {results1['avg_wait_time']:<10} {results2['avg_wait_time']:<12} {results3['avg_wait_time']:<12}")
print(f"Max Wait Time    {results1['max_wait_time']:<10} {results2['max_wait_time']:<12} {results3['max_wait_time']:<12}")
print(f"Queue Length     {results1['avg_queue_length']:<10} {results2['avg_queue_length']:<12} {results3['avg_queue_length']:<12}")
print(f"Throughput       {results1['throughput']:<10} {results2['throughput']:<12} {results3['throughput']:<12}")
print(f"Utilization      {results1['utilization']:<10} {results2['utilization']:<12} {results3['utilization']:<12}")

print("\n" + "=" * 70)
print("SIMULATION COMPLETE!")
print(f"Excel: {excel_file}")
print("=" * 70)